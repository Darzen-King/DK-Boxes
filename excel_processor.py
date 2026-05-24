from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os
import tempfile
from datetime import datetime

class ExcelProcessor:
    def __init__(self):
        self.thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        self.header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        self.header_font = Font(bold=True, color="FFFFFF", size=12)

    def read_products(self, file):
        """讀取 Excel 文件中的產品名稱"""
        products = []

        try:
            # 將上傳的文件加載為工作簿
            workbook = load_workbook(file)
            worksheet = workbook.active

            # 讀取第一列的產品名稱
            for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, values_only=True):
                product = row[0]
                if product and isinstance(product, str):
                    product = product.strip()
                    if product and product not in ['產品名稱', '商品名', 'Product', 'Name']:
                        products.append(product)

            return products
        except Exception as e:
            raise Exception(f"讀取 Excel 文件失敗: {str(e)}")

    def write_results(self, results_dict):
        """將搜尋結果寫入 Excel 文件"""
        try:
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "搜尋結果"

            # 設定欄寬
            worksheet.column_dimensions['A'].width = 20
            worksheet.column_dimensions['B'].width = 15
            worksheet.column_dimensions['C'].width = 50
            worksheet.column_dimensions['D'].width = 15
            worksheet.column_dimensions['E'].width = 15

            # 寫入表頭
            headers = ['產品名稱', '序號', '圖片 URL', '來源', '是否含水印']
            for col, header in enumerate(headers, 1):
                cell = worksheet.cell(row=1, column=col)
                cell.value = header
                cell.font = self.header_font
                cell.fill = self.header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = self.thin_border

            # 寫入數據
            row_idx = 2
            for product_name, product_data in results_dict.items():
                images = product_data.get('images', [])

                if not images:
                    # 即使沒有圖片也添加產品行
                    cell_a = worksheet.cell(row=row_idx, column=1)
                    cell_a.value = product_name
                    cell_a.border = self.thin_border
                    row_idx += 1
                else:
                    for idx, image_info in enumerate(images, 1):
                        cell_a = worksheet.cell(row=row_idx, column=1)
                        cell_a.value = product_name if idx == 1 else ""
                        cell_a.border = self.thin_border

                        cell_b = worksheet.cell(row=row_idx, column=2)
                        cell_b.value = idx
                        cell_b.alignment = Alignment(horizontal='center')
                        cell_b.border = self.thin_border

                        cell_c = worksheet.cell(row=row_idx, column=3)
                        cell_c.value = image_info.get('url', '')
                        cell_c.hyperlink = image_info.get('url', '')
                        cell_c.font = Font(color="0563C1", underline="single")
                        cell_c.border = self.thin_border

                        cell_d = worksheet.cell(row=row_idx, column=4)
                        cell_d.value = image_info.get('source', 'Unknown')
                        cell_d.alignment = Alignment(horizontal='center')
                        cell_d.border = self.thin_border

                        cell_e = worksheet.cell(row=row_idx, column=5)
                        has_watermark = "是" if image_info.get('has_watermark') else "否"
                        cell_e.value = has_watermark
                        cell_e.alignment = Alignment(horizontal='center')
                        cell_e.border = self.thin_border

                        row_idx += 1

            # 凍結表頭
            worksheet.freeze_panes = "A2"

            # 保存到臨時文件
            temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
            temp_file_path = temp_file.name
            temp_file.close()

            workbook.save(temp_file_path)
            return temp_file_path

        except Exception as e:
            raise Exception(f"寫入 Excel 文件失敗: {str(e)}")

    def create_template(self):
        """創建 Excel 模板文件"""
        try:
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "產品列表"

            # 設定欄寬
            worksheet.column_dimensions['A'].width = 30

            # 寫入表頭
            header_cell = worksheet.cell(row=1, column=1)
            header_cell.value = "產品名稱"
            header_cell.font = self.header_font
            header_cell.fill = self.header_fill
            header_cell.alignment = Alignment(horizontal='center', vertical='center')
            header_cell.border = self.thin_border

            # 添加示例
            example_products = ['Apple iPhone 13', 'Samsung Galaxy S21', 'Sony WH-1000XM4']
            for idx, product in enumerate(example_products, 2):
                cell = worksheet.cell(row=idx, column=1)
                cell.value = product
                cell.border = self.thin_border

            # 保存到臨時文件
            temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
            temp_file_path = temp_file.name
            temp_file.close()

            workbook.save(temp_file_path)
            return temp_file_path

        except Exception as e:
            raise Exception(f"創建模板失敗: {str(e)}")
